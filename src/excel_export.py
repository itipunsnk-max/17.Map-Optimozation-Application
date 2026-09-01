"""Formatted multi-sheet Excel export."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from .models import AnalysisResult


ROUTE_OUTPUT_COLUMNS = [
    "Branch_ID", "Branch_Name", "Province", "Input_Latitude", "Input_Longitude", "Resolved_Latitude", "Resolved_Longitude", "Reference_Latitude", "Reference_Longitude",
    "Location_Method", "Coordinate_Source", "Assigned_Hub_ID", "Assigned_Hub_Name", "Assigned_Region", "Hub_Latitude", "Hub_Longitude", "Hub_Location_Method", "Hub_Coordinate_Source",
    "Road_Distance_km", "Travel_Time_min", "Straight_Line_Distance_km", "Rank_2_Hub_ID", "Rank_2_Hub_Name", "Rank_2_Distance_km",
    "Rank_3_Hub_ID", "Rank_3_Hub_Name", "Rank_3_Distance_km", "Distance_Band", "Routing_Profile", "Routing_Source", "Calculation_Date", "Application_Version", "Status", "Analysis_Level",
]


def _public_route_results(frame: pd.DataFrame) -> pd.DataFrame:
    output = frame.copy()
    if "Analysis_Level" not in output.columns:
        output["Analysis_Level"] = "Branch"
    for column in ROUTE_OUTPUT_COLUMNS:
        if column not in output.columns:
            output[column] = ""
    return output[ROUTE_OUTPUT_COLUMNS]


def export_analysis_excel(result: AnalysisResult, path_or_buffer: str | Path | object) -> None:
    """Write all audit sheets to an xlsx path or file-like object."""
    route_results = _public_route_results(result.route_results)
    sheets = {
        "Route_Results": route_results,
        "Distance_Matrix": result.distance_matrix,
        "Duration_Matrix": result.duration_matrix,
        "Province_Summary": result.province_summary,
        "Hub_Summary": result.hub_summary,
        "Validation_Errors": result.validation_errors,
        "Failed_Routes": result.failed_routes,
    }
    with pd.ExcelWriter(path_or_buffer, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            frame = frame if isinstance(frame, pd.DataFrame) else pd.DataFrame(frame)
            frame.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            for column_cells in worksheet.columns:
                values = [str(cell.value) if cell.value is not None else "" for cell in column_cells[:100]]
                width = min(max(max((len(value) for value in values), default=10) + 2, 10), 36)
                worksheet.column_dimensions[column_cells[0].column_letter].width = width
            if worksheet.max_row > 1 and worksheet.max_column > 0:
                ref = f"A1:{worksheet.cell(worksheet.max_row, worksheet.max_column).coordinate}"
                table = Table(displayName=f"Tbl{sheet_index}{worksheet.title.replace('_', '')}", ref=ref)
                table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
                worksheet.add_table(table)
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, float):
                        cell.number_format = "0.00"
            for column_name in ("Road_Distance_km", "Average_Road_Distance_km", "Maximum_Road_Distance_km", "Distance_km"):
                if column_name in [cell.value for cell in worksheet[1]] and worksheet.max_row > 1:
                    column_index = [cell.value for cell in worksheet[1]].index(column_name) + 1
                    letter = worksheet.cell(1, column_index).column_letter
                    worksheet.conditional_formatting.add(f"{letter}2:{letter}{worksheet.max_row}", ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="F8696B"))
