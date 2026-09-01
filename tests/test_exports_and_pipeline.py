from dataclasses import replace
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from config.settings import get_settings
from src.excel_export import export_analysis_excel
from src.excel_loader import load_workbook as load_input_workbook
from src.geojson_export import build_geojson
from src.pipeline import run_analysis
from src.routing_provider import OfflineRoutingProvider


def _input_frames():
    branches = pd.DataFrame([
        {"Branch_ID": "B001", "Branch_Name": "Rayong", "Province": "ระยอง", "Latitude": 12.681, "Longitude": 101.281},
        {"Branch_ID": "B002", "Branch_Name": "Chiang Mai", "Province": "เชียงใหม่", "Latitude": "", "Longitude": ""},
    ])
    hubs = pd.DataFrame([
        {"Hub_ID": "H1", "Region": "Central", "Hub_Name": "Bangkok Hub", "Province": "กรุงเทพมหานคร", "Latitude": 13.7563, "Longitude": 100.5018},
        {"Hub_ID": "H2", "Region": "North", "Hub_Name": "Chiang Mai Hub", "Province": "เชียงใหม่", "Latitude": 18.7883, "Longitude": 98.9853},
        {"Hub_ID": "H3", "Region": "East", "Hub_Name": "Rayong Hub", "Province": "ระยอง", "Latitude": 12.6825, "Longitude": 101.275},
    ])
    return branches, hubs


def test_pipeline_and_all_exports(tmp_path):
    settings = replace(get_settings(), cache_path=tmp_path / "routing.sqlite3", output_dir=tmp_path, log_dir=tmp_path / "logs")
    branches, hubs = _input_frames()
    result = run_analysis(branches, hubs, OfflineRoutingProvider(), settings=settings, force_recalculate=True)
    assert len(result.route_results) == 2
    assert set(result.route_results["Status"]) == {"OK"}
    assert set(result.distance_matrix.columns) == {"Branch_ID", "Branch_Name", "Distance_to_H1_km", "Distance_to_H2_km", "Distance_to_H3_km"}
    assert "Duration_to_H1_min" in result.duration_matrix.columns
    assert not result.province_summary.empty
    assert len(result.hub_summary) == 3
    assert len(build_geojson(result.route_results)["features"]) == 2
    output = tmp_path / "route_results.xlsx"
    export_analysis_excel(result, output)
    book = load_workbook(output, read_only=True)
    assert set(book.sheetnames) == {"Route_Results", "Distance_Matrix", "Duration_Matrix", "Province_Summary", "Hub_Summary", "Validation_Errors", "Failed_Routes"}
    assert book["Route_Results"].max_row == 3


def test_excel_loader_round_trip(tmp_path):
    branches, hubs = _input_frames()
    source = tmp_path / "input.xlsx"
    with pd.ExcelWriter(source, engine="openpyxl") as writer:
        branches.to_excel(writer, sheet_name="Branches", index=False)
        hubs.to_excel(writer, sheet_name="Regional_Hubs", index=False)
    loaded_branches, loaded_hubs = load_input_workbook(source)
    assert list(loaded_branches.columns) == list(branches.columns)
    assert len(loaded_hubs) == 3


def test_province_level_and_assigned_geometry_only(tmp_path):
    class CountingProvider(OfflineRoutingProvider):
        def __init__(self):
            super().__init__()
            self.matrix_calls = 0
            self.geometry_calls = 0

        def calculate_matrix(self, origins, destinations, profile):
            self.matrix_calls += 1
            return super().calculate_matrix(origins, destinations, profile)

        def get_route(self, origin, destination, profile):
            self.geometry_calls += 1
            return super().get_route(origin, destination, profile)

    settings = replace(get_settings(), cache_path=tmp_path / "routing.sqlite3", matrix_batch_size=1)
    branches, hubs = _input_frames()
    provider = CountingProvider()
    result = run_analysis(branches, hubs, provider, settings=settings, analysis_level="Province Level", force_recalculate=True)
    assert set(result.route_results["Analysis_Level"]) == {"Province"}
    assert result.route_results["Reference_Latitude"].notna().all()
    assert provider.matrix_calls == len(result.route_results)
    assert provider.geometry_calls == len(result.route_results)
